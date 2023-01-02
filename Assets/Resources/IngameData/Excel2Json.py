# Need to install openpyxl
# pip install openpyxl
# python .\Assets\Scripts\excel_to_json.py path to file
# Example: python .\Assets\Scripts\excel_to_json.py Assets data UnitData EnemyData

from openpyxl import load_workbook
import json
import os

SAVE_FILE_PATH = 'Json'
ROW_INDEX = 'B1'
COLUMN_INDEX = 'C1'
START_ALPHABET = 'B'
SELECTING_KEY = ',,,\n'
KEY_START_NUM = 3
DATA_START_NUM = 5
AV_DATA_TYPE_STR = ["int", "string", "bool", "int[]", "string[]", "bool[]"]

def getData(typestr, data):
    if typestr == "int":
        return int(data)
    if typestr == "string":
        return str(data)
    if typestr == "bool":
        if data == "true" or data == "1" or data == "True" or data == "TRUE" or data == 1:
            return True
        if data == "False" or data == "0" or data == "false" or data == "FALSE":
            return False
        else:
            print("Data \"" + str(data) + "\" is converted to \"False\"")
            return False
    if typestr == "int[]":
        return list(map(lambda x: getData("int", x), str(data).split(SELECTING_KEY)))
    if typestr == "string[]":
        return list(map(lambda x: getData("string", str(x)), str(data).split(SELECTING_KEY)))
    if typestr == "bool[]":
        return list(map(lambda x: getData("bool", x), str(data).split(SELECTING_KEY)))

SAVE_FILE_PATH = os.path.join(os.getcwd(), SAVE_FILE_PATH)
file_path = os.getcwd()
file_path = os.path.join(file_path, 'DataBase')

excel_file_path =  file_path + '.xlsx'

wb = load_workbook(excel_file_path, data_only=True)
workbook_num = len(wb.worksheets)

if not os.path.exists(SAVE_FILE_PATH):
    os.makedirs(SAVE_FILE_PATH)

for sheetidx in range(workbook_num):
    worksheet = wb.worksheets[sheetidx]
    sheetname = wb.sheetnames[sheetidx]
    data_size = int(worksheet[ROW_INDEX].value)
    datatype_len = int(worksheet[COLUMN_INDEX].value)

    # Get Datas
    
    datatype_list = []
    for i in range(datatype_len):
        curr_datatype_key = str(chr(ord(START_ALPHABET) + i)) + str(KEY_START_NUM)
        curr_datatype = worksheet[curr_datatype_key].value
        assert (curr_datatype in AV_DATA_TYPE_STR), "\n\nINPUT DATA TYPE ERROR!\n" + "Worksheet name: " + wb.sheetnames[sheetidx] + "\nPosition: " + curr_datatype_key + "\n" + "\"" + curr_datatype + "\" is not avaliable input type\nAllowed Data Type:" + str(AV_DATA_TYPE_STR)
        datatype_list.append(curr_datatype)
    print(datatype_list)

    key_list = []
    for i in range(datatype_len):
        curr_datatype_key = str(chr(ord(START_ALPHABET) + i)) + str(KEY_START_NUM + 1)
        curr_datatype = worksheet[curr_datatype_key].value
        key_list.append(curr_datatype)

    data_list = []
    for i in range(data_size):
        data_list_list = []
        for datatype_index in range(datatype_len):
            curr_datatype_key = str(chr(ord(START_ALPHABET) + datatype_index)) + str(DATA_START_NUM + i)
            data_list_list.append(getData(datatype_list[datatype_index], worksheet[curr_datatype_key].value))
        data_list.append(data_list_list)
    print(data_list)

    json_data_list = []
    for data_idx in range(data_size):
        idx_dict = {}
        datas = data_list[data_idx]
        for key_idx in range(datatype_len):
            idx_dict[key_list[key_idx]] = datas[key_idx]
        json_data_list.append(idx_dict)

    json_file_path = os.path.join(SAVE_FILE_PATH, wb.sheetnames[sheetidx] + '.json')
    with open(json_file_path, 'w', encoding='UTF8') as fd:
        json.dump(json_data_list, fd, sort_keys=True, indent=4)

os.system("pause")